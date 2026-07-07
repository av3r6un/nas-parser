"""Tests for exported catalog audit diagnostics."""

from __future__ import annotations

from contextlib import redirect_stdout
from io import StringIO
from pathlib import Path
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from nas_parser.audit import CatalogAudit, main


def _write_catalog_workbook(workbook_path: Path) -> None:
    """Create a small exported catalog workbook with audit issues."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары"
    sheet.append(
        (
            "Артикул",
            "Наименование",
            "Цена",
            "Количество",
            "Категория",
            "Цвет",
            "Размер",
            "Форма",
            "Фиксация",
            "Грани",
        )
    )
    sheet.append(
        (
            None,
            "Crystal",
            None,
            5,
            "Стразы горячей фиксации",
            "001",
            "SS16",
            None,
            "Hot",
            "12cut",
        )
    )
    sheet.append(
        (
            "DUP-1",
            "Crystal AB",
            -10,
            None,
            "Стразы холодной фиксации",
            "001+",
            "SS30",
            None,
            "Non",
            "16cut",
        )
    )
    sheet.append(
        (
            "DUP-1",
            "Drop",
            25,
            7,
            "Пришивные стразы",
            "001",
            "7*12",
            "Drop",
            "K9",
            None,
        )
    )
    workbook.save(workbook_path)
    workbook.close()


def _write_clean_catalog_workbook(workbook_path: Path) -> None:
    """Create a small exported catalog workbook without audit issues."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары"
    sheet.append(
        (
            "Артикул",
            "Наименование",
            "Цена",
            "Количество",
            "Категория",
            "Цвет",
            "Размер",
            "Форма",
            "Фиксация",
            "Грани",
        )
    )
    sheet.append(
        (
            "12cut/Crystal/SS16/Hot/001",
            "Crystal",
            100,
            5,
            "Стразы горячей фиксации",
            "001",
            "SS16",
            None,
            "Hot",
            "12cut",
        )
    )
    sheet.append(
        (
            "Crystal/Drop/10x14",
            "Crystal Drop",
            25,
            7,
            "Пришивные стразы",
            "001",
            "10x14",
            "Drop",
            "K9",
            None,
        )
    )
    workbook.save(workbook_path)
    workbook.close()


class TestCatalogAudit(unittest.TestCase):
    """Coverage for the exported catalog audit tool."""

    def test_run_detects_catalog_quality_issues(self) -> None:
        """Verify the audit detects empty and problematic exported values."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "catalog.xlsx"
            _write_catalog_workbook(workbook_path)

            report = CatalogAudit(workbook_path).run()

        self.assertEqual(report.rows, 3)
        self.assertEqual(report.empty_sku, 1)
        self.assertEqual(report.duplicate_sku, 1)
        self.assertEqual(report.empty_name, 0)
        self.assertEqual(report.empty_price, 1)
        self.assertEqual(report.negative_price, 1)
        self.assertEqual(report.empty_quantity, 1)
        self.assertEqual(report.negative_quantity, 0)
        self.assertEqual(report.formula_quantity, 0)
        self.assertEqual(report.empty_shape, 2)
        self.assertEqual(report.empty_fixation, 0)
        self.assertEqual(report.empty_cut, 1)
        self.assertEqual(report.empty_sku_rows, [2])
        self.assertEqual(report.empty_price_rows, [2])
        self.assertEqual(report.empty_quantity_rows, [3])
        self.assertEqual(report.duplicate_sku_rows, {"DUP-1": [3, 4]})
        self.assertEqual(
            report.category_counts,
            {
                "Пришивные стразы": 1,
                "Стразы горячей фиксации": 1,
                "Стразы холодной фиксации": 1,
            },
        )

    def test_summary_formats_statistics_and_problem_rows(self) -> None:
        """Verify the audit summary includes detailed diagnostics."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "catalog.xlsx"
            _write_catalog_workbook(workbook_path)

            summary = CatalogAudit(workbook_path).run().summary()

        self.assertIn("Rows: 3", summary)
        self.assertIn("SKU:", summary)
        self.assertIn("  empty: 1", summary)
        self.assertIn("  duplicates: 1", summary)
        self.assertIn("Price:", summary)
        self.assertIn("  empty: 1", summary)
        self.assertIn("  negative: 1", summary)
        self.assertIn("Quantity:", summary)
        self.assertIn("  empty: 1", summary)
        self.assertIn("Category:", summary)
        self.assertIn("Стразы горячей фиксации: 1", summary)
        self.assertIn("Duplicate SKU:", summary)
        self.assertIn("DUP-1", summary)
        self.assertIn("Rows:\n3\n4", summary)
        self.assertIn("Empty SKU:\nRows:\n2", summary)
        self.assertIn("Empty Price:\nRows:\n2", summary)
        self.assertIn("Empty Quantity:\nRows:\n3", summary)

    def test_summary_omits_detail_sections_when_no_problems_exist(self) -> None:
        """Verify the audit does not print extra diagnostics for a clean catalog."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "catalog.xlsx"
            _write_clean_catalog_workbook(workbook_path)

            summary = CatalogAudit(workbook_path).run().summary()

        self.assertIn("Rows: 2", summary)
        self.assertIn("  empty: 0", summary)
        self.assertIn("  duplicates: 0", summary)
        self.assertNotIn("Duplicate SKU:", summary)
        self.assertNotIn("Empty SKU:", summary)
        self.assertNotIn("Empty Price:", summary)
        self.assertNotIn("Empty Quantity:", summary)

    def test_audit_accepts_generated_k9_sku_without_false_issues(self) -> None:
        """Verify generated K9 SKU values are treated as normal non-duplicate SKUs."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "catalog.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Товары"
            sheet.append(
                (
                    "Артикул",
                    "Наименование",
                    "Цена",
                    "Количество",
                    "Категория",
                    "Цвет",
                    "Размер",
                    "Форма",
                    "Фиксация",
                    "Грани",
                )
            )
            sheet.append(
                (
                    "K9/Aurora Green/Navette/7x15/203",
                    "Aurora Green Navette",
                    11,
                    1886,
                    "Пришивные стразы",
                    "203",
                    "7x15",
                    "Navette",
                    "K9",
                    None,
                )
            )
            sheet.append(
                (
                    "K9/Light Rose/Rivoli/12x12/204",
                    "Light Rose Rivoli",
                    12,
                    100,
                    "Пришивные стразы",
                    "204",
                    "12x12",
                    "Rivoli",
                    "K9",
                    None,
                )
            )
            workbook.save(workbook_path)
            workbook.close()

            report = CatalogAudit(workbook_path).run()

        self.assertEqual(report.empty_sku, 0)
        self.assertEqual(report.duplicate_sku, 0)
        self.assertEqual(report.empty_sku_rows, [])
        self.assertEqual(report.duplicate_sku_rows, {})


class TestCatalogAuditCli(unittest.TestCase):
    """Coverage for the catalog audit command-line entry point."""

    def test_main_returns_error_when_argument_is_missing(self) -> None:
        """Verify the CLI prints usage when the workbook path is omitted."""
        stdout = StringIO()

        with patch("sys.argv", ["python"]), redirect_stdout(stdout):
            exit_code = main()

        self.assertEqual(exit_code, 1)
        self.assertIn("Usage:", stdout.getvalue())
        self.assertIn("python -m nas_parser.audit output/catalog.xlsx", stdout.getvalue())

    def test_main_prints_summary_for_existing_workbook(self) -> None:
        """Verify the CLI runs the audit and prints its summary."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "catalog.xlsx"
            _write_catalog_workbook(workbook_path)
            stdout = StringIO()

            with patch("sys.argv", ["python", str(workbook_path)]), redirect_stdout(stdout):
                exit_code = main()

        self.assertEqual(exit_code, 0)
        self.assertIn("Rows: 3", stdout.getvalue())
        self.assertIn("Duplicate SKU:", stdout.getvalue())

    def test_main_returns_error_for_missing_workbook(self) -> None:
        """Verify the CLI prints a readable message for a missing workbook."""
        stdout = StringIO()
        missing_path = Path("output/missing-catalog.xlsx")

        with patch("sys.argv", ["python", str(missing_path)]), redirect_stdout(stdout):
            exit_code = main()

        self.assertEqual(exit_code, 1)
        self.assertIn(f"File not found: {missing_path}", stdout.getvalue())
