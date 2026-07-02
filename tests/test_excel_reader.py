"""Tests for NAS Parser Excel reader implementation."""

from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook

from nas_parser.readers import ExcelReader
from nas_parser.source import SourceRow


class TestExcelReader(unittest.TestCase):
    """Coverage for the ExcelReader implementation."""

    def test_excel_reader_reads_rows_from_a_workbook(self) -> None:
        """Verify that ExcelReader yields SourceRow objects from a workbook."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "sample.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(["Topaz", 12, 3.5])
            sheet.append([None, None, None])
            sheet.append(["AB", "SS16", "Hotfix"])
            other_sheet = workbook.create_sheet("Sheet2")
            other_sheet.append(["K9", "Rivoli"])
            workbook.save(workbook_path)
            workbook.close()

            reader = ExcelReader(workbook_path)
            rows = list(reader.read())

        self.assertEqual(reader.name, "excel")
        self.assertEqual(len(rows), 3)
        self.assertIsInstance(rows[0], SourceRow)
        self.assertEqual(rows[0].source_file, workbook_path)
        self.assertEqual(rows[0].source_sheet, "Sheet1")
        self.assertEqual(rows[0].source_row, 1)
        self.assertEqual(rows[0].values, ("Topaz", 12, 3.5))
        self.assertEqual(rows[1].source_row, 3)
        self.assertEqual(rows[1].values, ("AB", "SS16", "Hotfix"))
        self.assertEqual(rows[2].source_sheet, "Sheet2")
        self.assertEqual(rows[2].values, ("K9", "Rivoli"))

    def test_excel_reader_skips_fully_empty_rows(self) -> None:
        """Verify that fully empty rows are skipped."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "empty_rows.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            sheet.append([None, None])
            sheet.append(["Value", None])
            workbook.save(workbook_path)
            workbook.close()

            reader = ExcelReader(workbook_path)
            rows = list(reader.read())

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].values, ("Value", None))

    def test_excel_reader_preserves_real_excel_row_numbers(self) -> None:
        """Verify that empty rows are skipped while real Excel row numbers remain intact."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "row_numbers.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            sheet.append(["First"])
            sheet.append([None])
            sheet.append(["Third"])
            workbook.save(workbook_path)
            workbook.close()

            reader = ExcelReader(workbook_path)
            rows = list(reader.read())

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].source_row, 1)
        self.assertEqual(rows[0].values, ("First",))
        self.assertEqual(rows[1].source_row, 3)
        self.assertEqual(rows[1].values, ("Third",))

    def test_excel_reader_keeps_original_value_types(self) -> None:
        """Verify that ExcelReader does not normalize or coerce cell value types."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "types.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Types"
            sheet.append(["Text", 7, 3.5, True, None])
            workbook.save(workbook_path)
            workbook.close()

            reader = ExcelReader(workbook_path)
            rows = list(reader.read())

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].values, ("Text", 7, 3.5, True, None))
        self.assertIsInstance(rows[0].values[0], str)
        self.assertIsInstance(rows[0].values[1], int)
        self.assertIsInstance(rows[0].values[2], float)
        self.assertIsInstance(rows[0].values[3], bool)
        self.assertIsNone(rows[0].values[4])
