"""Tests for NAS Parser reference infrastructure."""

from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from nas_parser.references import ColorReference, ColorReferenceManager, ReferenceBase, ReferenceLoader
from nas_parser.references import ColorReferenceLoader
from nas_parser.references.manager import GeneratedColorRecord


def _write_manager_reference_workbook(workbook_path: Path) -> None:
    """Create a color reference workbook for manager tests."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(
        [
            "Артикул WB и Ozon",
            "Грани",
            "Цвет",
            "Размер",
            "Тип",
            "Артикул цвета",
        ]
    )
    sheet.append(["12cut/Crystal/SS3/Hot/001", "12cut", "Crystal", "SS3", "Hot", "001"])
    sheet.append(["12cut/Topaz/SS3/Hot/008", "12cut", "Topaz", "SS3", "Hot", "008"])
    workbook.save(workbook_path)
    workbook.close()


class TestColorReference(unittest.TestCase):
    """Coverage for the color reference contract."""

    def test_color_reference_exposes_the_expected_contract(self) -> None:
        """Verify that ColorReference exposes the base reference fields."""
        reference = ColorReference(source_file=Path("reference/colorcode-articul.xlsx"))

        self.assertEqual(reference.name, "colors")
        self.assertEqual(reference.source_file, Path("reference/colorcode-articul.xlsx"))
        self.assertFalse(reference.loaded)
        self.assertIsInstance(reference, ReferenceBase)


class TestReferenceLoader(unittest.TestCase):
    """Coverage for the reference loader infrastructure."""

    def test_reference_loader_registers_and_returns_references(self) -> None:
        """Verify that the loader stores and retrieves references by name."""
        reference = ColorReference()
        loader = ReferenceLoader()

        loader.register(reference)

        self.assertIs(loader.get("colors"), reference)
        self.assertEqual(loader.all(), (reference,))

    def test_reference_loader_returns_none_for_unknown_reference(self) -> None:
        """Verify that the loader returns None when a name is not registered."""
        loader = ReferenceLoader()

        self.assertIsNone(loader.get("missing"))


class TestColorReferenceLoader(unittest.TestCase):
    """Coverage for loading color codes from an Excel workbook."""

    def test_color_reference_loader_loads_codes_into_lookup_structure(self) -> None:
        """Verify that color names are loaded into a fast lookup structure."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "colorcode-articul.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Лист1"
            sheet.append(["Артикул WB и Ozon", "Грани", "Цвет", "Размер", "Тип", "Артикул цвета"])
            sheet.append(["12cut/Crystal/SS3/Hot/001", "12cut", "Crystal", "SS3", "Hot", "001"])
            sheet.append(["12cut/Topaz/SS3/Hot/008", "12cut", "Topaz", "SS3", "Hot", "008"])
            sheet.append([None, None, None, None, None, None])
            workbook.save(workbook_path)
            workbook.close()

            loader = ColorReferenceLoader(workbook_path)
            reference = loader.load()

        self.assertTrue(reference.loaded)
        self.assertEqual(reference.source_file, workbook_path)
        self.assertEqual(reference.get_code("Crystal"), "001")
        self.assertEqual(reference.get_code("Topaz"), "008")
        self.assertEqual(reference.get_code("Unknown"), None)

    def test_color_reference_loader_strips_outer_whitespace_only(self) -> None:
        """Verify that lookup handles surrounding whitespace without further normalization."""
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "colorcode-articul.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Лист1"
            sheet.append(["Артикул WB и Ozon", "Грани", "Цвет", "Размер", "Тип", "Артикул цвета"])
            sheet.append(["12cut/Crystal/SS3/Hot/001", "12cut", "Crystal", "SS3", "Hot", "001"])
            workbook.save(workbook_path)
            workbook.close()

            reference = ColorReferenceLoader(workbook_path).load()

        self.assertEqual(reference.get_code("  Crystal  "), "001")


class TestColorReferenceManager(unittest.TestCase):
    """Coverage for managing generated color reference files."""

    def test_active_reference_uses_base_file_when_generated_is_absent(self) -> None:
        """Verify the base reference is active when no generated directory exists."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            base_file = reference_dir / "colorcode-articul.xlsx"
            base_file.write_bytes(b"base")

            manager = ColorReferenceManager(reference_dir)

            self.assertEqual(manager.get_active_reference_file(), base_file)

    def test_active_reference_uses_gen1_when_it_exists(self) -> None:
        """Verify gen1 is active when it is the only generated reference."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            generated_dir = reference_dir / "generated"
            generated_dir.mkdir()
            gen1 = generated_dir / "colorcode-articul_gen1.xlsx"
            gen1.write_bytes(b"gen1")

            manager = ColorReferenceManager(reference_dir)

            self.assertEqual(manager.get_active_reference_file(), gen1)

    def test_active_reference_uses_highest_generation_file(self) -> None:
        """Verify the active reference is the generated file with the max generation."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            generated_dir = reference_dir / "generated"
            generated_dir.mkdir()
            (generated_dir / "colorcode-articul_gen1.xlsx").write_bytes(b"gen1")
            (generated_dir / "colorcode-articul_gen2.xlsx").write_bytes(b"gen2")
            gen5 = generated_dir / "colorcode-articul_gen5.xlsx"
            gen5.write_bytes(b"gen5")

            manager = ColorReferenceManager(reference_dir)

            self.assertEqual(manager.get_active_reference_file(), gen5)

    def test_next_generation_file_returns_next_available_generation(self) -> None:
        """Verify the next generation path is computed from the max existing generation."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            generated_dir = reference_dir / "generated"
            generated_dir.mkdir()
            (generated_dir / "colorcode-articul_gen1.xlsx").write_bytes(b"gen1")
            (generated_dir / "colorcode-articul_gen2.xlsx").write_bytes(b"gen2")
            (generated_dir / "colorcode-articul_gen5.xlsx").write_bytes(b"gen5")
            (generated_dir / "colorcode-articul_gen7.xlsx").write_bytes(b"gen7")

            manager = ColorReferenceManager(reference_dir)

            self.assertEqual(
                manager.next_generation_file(),
                generated_dir / "colorcode-articul_gen8.xlsx",
            )

    def test_next_generation_file_creates_generated_directory(self) -> None:
        """Verify the generated directory is created when computing the first generation."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            generated_dir = reference_dir / "generated"

            manager = ColorReferenceManager(reference_dir)
            next_file = manager.next_generation_file()

            self.assertTrue(generated_dir.is_dir())
            self.assertEqual(next_file, generated_dir / "colorcode-articul_gen1.xlsx")
            self.assertFalse(next_file.exists())

    def test_copy_reference_for_generation_creates_copy_without_changing_source(self) -> None:
        """Verify generation copies preserve the active source file unchanged."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            base_file = reference_dir / "colorcode-articul.xlsx"
            base_file.write_bytes(b"base reference content")

            manager = ColorReferenceManager(reference_dir)
            generated_file = manager.copy_reference_for_generation()

            self.assertEqual(generated_file, reference_dir / "generated" / "colorcode-articul_gen1.xlsx")
            self.assertTrue(generated_file.is_file())
            self.assertEqual(generated_file.read_bytes(), b"base reference content")
            self.assertEqual(base_file.read_bytes(), b"base reference content")

    def test_write_generated_records_appends_rows_to_generation_file(self) -> None:
        """Verify generated records are appended to the copied reference workbook."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            base_file = reference_dir / "colorcode-articul.xlsx"
            _write_manager_reference_workbook(base_file)
            base_bytes = base_file.read_bytes()

            manager = ColorReferenceManager(reference_dir)
            manager.ensure_color("Aurora Green", "16cut", "SS16", "non")
            generation_file = manager.copy_reference_for_generation()
            manager.write_generated_records(generation_file)

            generated_workbook = load_workbook(generation_file, read_only=True, data_only=True)
            try:
                rows = list(generated_workbook.active.iter_rows(values_only=True))
            finally:
                generated_workbook.close()
            unchanged_base_bytes = base_file.read_bytes()

        self.assertEqual(unchanged_base_bytes, base_bytes)
        self.assertIn(("16cut/Aurora Green/SS16/Non/203", "16cut", "Aurora Green", "SS16", "Non", "203"), rows)
        self.assertGreaterEqual(len(rows), 103)

    def test_write_generated_records_does_not_add_second_blank_block(self) -> None:
        """Verify later generations append records without another empty block."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            first_manager = ColorReferenceManager(reference_dir)
            first_manager.ensure_color("Aurora Green", "16cut", "SS16", "non")
            gen1 = first_manager.copy_reference_for_generation()
            first_manager.write_generated_records(gen1)

            second_manager = ColorReferenceManager(reference_dir)
            second_manager.ensure_color("Light Rose", "12cut", "SS20", "hot")
            gen2 = second_manager.copy_reference_for_generation()
            second_manager.write_generated_records(gen2)

            workbook = load_workbook(gen2, read_only=True, data_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()

        non_empty_rows = [row for row in rows if any(value is not None for value in row)]
        self.assertEqual(non_empty_rows[-2][2], "Aurora Green")
        self.assertEqual(non_empty_rows[-1][2], "Light Rose")

    def test_write_generated_records_skips_existing_normalized_color(self) -> None:
        """Verify repeated writes do not duplicate generated color rows."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            manager.ensure_color("Aurora Green", "16cut", "SS16", "non")
            generation_file = manager.copy_reference_for_generation()
            manager.write_generated_records(generation_file)
            manager.write_generated_records(generation_file)

            workbook = load_workbook(generation_file, read_only=True, data_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()

        color_rows = [row for row in rows if row[2] == "Aurora Green"]
        self.assertEqual(len(color_rows), 1)
        self.assertEqual(color_rows[0], ("16cut/Aurora Green/SS16/Non/203", "16cut", "Aurora Green", "SS16", "Non", "203"))

    def test_ensure_color_returns_existing_color_code(self) -> None:
        """Verify existing reference colors are returned without generation."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            color_code = manager.ensure_color("Crystal", "12cut", "SS16", "non")

        self.assertEqual(color_code, "001")
        self.assertEqual(manager.generated_records(), ())

    def test_ensure_color_normalizes_existing_reference_color_lookup(self) -> None:
        """Verify existing colors are matched using normalized color identity."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            codes = [
                manager.ensure_color(" crystal ", "12cut", "SS16", "non"),
                manager.ensure_color("CRYSTAL", "12cut", "SS16", "non"),
                manager.ensure_color("Crystal\u00a0", "12cut", "SS16", "non"),
                manager.ensure_color("Ｃｒｙｓｔａｌ", "12cut", "SS16", "non"),
            ]

        self.assertEqual(codes, ["001", "001", "001", "001"])
        self.assertEqual(manager.generated_records(), ())

    def test_ensure_color_creates_new_color_record(self) -> None:
        """Verify a missing color creates an in-memory generated record."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            color_code = manager.ensure_color("Aurora Green", "16cut", "SS16", "non")
            records = manager.generated_records()

        self.assertEqual(color_code, "203")
        self.assertEqual(
            records,
            (
                GeneratedColorRecord(
                    article="16cut/Aurora Green/SS16/Non/203",
                    cut="16cut",
                    color="Aurora Green",
                    size="SS16",
                    type="Non",
                    color_code="203",
                ),
            ),
        )

    def test_ensure_color_generates_sequential_codes_for_multiple_new_colors(self) -> None:
        """Verify new generated color codes increment from 203."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            first_code = manager.ensure_color("Aurora Green", "16cut", "SS16", "non")
            second_code = manager.ensure_color("Light Rose", "12cut", "SS20", "hot")
            third_code = manager.ensure_color("Montana", "Navette", "7x15", "sew")

        self.assertEqual(first_code, "203")
        self.assertEqual(second_code, "204")
        self.assertEqual(third_code, "205")
        self.assertEqual([record.color_code for record in manager.generated_records()], ["203", "204", "205"])
        self.assertEqual(manager.generated_records()[2].article, "K9/Montana/Navette/7x15/205")

    def test_ensure_color_uses_max_existing_code_without_reusing_gaps(self) -> None:
        """Verify generated codes continue from the max existing color code."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            workbook_path = reference_dir / "colorcode-articul.xlsx"
            _write_manager_reference_workbook(workbook_path)
            workbook = load_workbook(workbook_path)
            try:
                sheet = workbook.active
                sheet.append(["12cut/Color203/SS3/Hot/203", "12cut", "Color203", "SS3", "Hot", "203"])
                sheet.append(["12cut/Color204/SS3/Hot/204", "12cut", "Color204", "SS3", "Hot", "204"])
                sheet.append(["12cut/Color205/SS3/Hot/205", "12cut", "Color205", "SS3", "Hot", "205"])
                sheet.append(["12cut/Color209/SS3/Hot/209", "12cut", "Color209", "SS3", "Hot", "209"])
                workbook.save(workbook_path)
            finally:
                workbook.close()

            manager = ColorReferenceManager(reference_dir)
            color_code = manager.ensure_color("Aurora Green", "16cut", "SS16", "non")

        self.assertEqual(color_code, "210")

    def test_ensure_color_reuses_generated_color_code(self) -> None:
        """Verify repeated calls for the same generated color do not create records."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            first_code = manager.ensure_color("Aurora Green", "16cut", "SS16", "non")
            second_code = manager.ensure_color("Aurora Green", "12cut", "SS20", "hot")

        self.assertEqual(first_code, "203")
        self.assertEqual(second_code, "203")
        self.assertEqual(len(manager.generated_records()), 1)

    def test_ensure_color_normalizes_generated_color_identity(self) -> None:
        """Verify formatted variants of a new color reuse one generated record."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            codes = [
                manager.ensure_color("Aurora Green", "16cut", "SS16", "non"),
                manager.ensure_color("aurora green", "16cut", "SS16", "non"),
                manager.ensure_color("AURORA GREEN", "16cut", "SS16", "non"),
                manager.ensure_color(" Aurora Green ", "16cut", "SS16", "non"),
                manager.ensure_color("Aurora  Green", "16cut", "SS16", "non"),
                manager.ensure_color("Aurora\u00a0Green", "16cut", "SS16", "non"),
                manager.ensure_color("Ａｕｒｏｒａ\u3000Ｇｒｅｅｎ", "16cut", "SS16", "non"),
            ]
            records = manager.generated_records()

        self.assertEqual(codes, ["203", "203", "203", "203", "203", "203", "203"])
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].color, "Aurora Green")
        self.assertEqual(records[0].article, "16cut/Aurora Green/SS16/Non/203")

    def test_ensure_color_normalizes_k9_shape_display_to_title_case(self) -> None:
        """Verify generated K9 shapes are stored in a unified Title Case display."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            manager.ensure_color("Cosmic", "DROP", "7x12", "sew")
            manager.ensure_color("Nebula", "DIAMOND LEAF", "7x12", "sew")
            manager.ensure_color("Aurora", "COSMIC BAGUETTE", "7x12", "sew")
            records = manager.generated_records()

        self.assertEqual([record.cut for record in records], ["Drop", "Diamond Leaf", "Cosmic Baguette"])
        self.assertEqual(
            [record.article for record in records],
            [
                "K9/Cosmic/Drop/7x12/203",
                "K9/Nebula/Diamond Leaf/7x12/204",
                "K9/Aurora/Cosmic Baguette/7x12/205",
            ],
        )

    def test_write_generated_records_formats_color_code_display(self) -> None:
        """Verify generated reference rows store display codes with zero padding."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            manager._generated_records.append(
                GeneratedColorRecord(
                    article="K9/Aurora Green/Drop/7x12/052+",
                    cut="Drop",
                    color="Aurora Green",
                    size="7x12",
                    type="K9",
                    color_code="52+",
                )
            )
            generation_file = manager.copy_reference_for_generation()
            manager.write_generated_records(generation_file)

            workbook = load_workbook(generation_file, read_only=True, data_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()

        self.assertIn(
            ("K9/Aurora Green/Drop/7x12/052+", "K9", "Aurora Green", "7x12", "K9", "052+"),
            rows,
        )

    def test_format_color_code_display_uses_three_digits_below_100(self) -> None:
        """Verify the color code display formatter pads only small numeric codes."""
        self.assertEqual(ColorReferenceManager._format_color_code_display("1"), "001")
        self.assertEqual(ColorReferenceManager._format_color_code_display("16"), "016")
        self.assertEqual(ColorReferenceManager._format_color_code_display("52+"), "052+")
        self.assertEqual(ColorReferenceManager._format_color_code_display("203"), "203")
        self.assertEqual(ColorReferenceManager._format_color_code_display("203+"), "203+")

    def test_ensure_color_ab_uses_base_color_code_with_plus(self) -> None:
        """Verify AB variants use the base color numeric code with a plus suffix."""
        with tempfile.TemporaryDirectory() as temp_dir:
            reference_dir = Path(temp_dir)
            _write_manager_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            manager = ColorReferenceManager(reference_dir)
            base_code = manager.ensure_color("Aurora Green", "16cut", "SS16", "non")
            ab_code = manager.ensure_color("Aurora Green AB", "16cut", "SS16", "non")

        self.assertEqual(base_code, "203")
        self.assertEqual(ab_code, "203+")
        self.assertEqual([record.color_code for record in manager.generated_records()], ["203", "203+"])
        self.assertEqual(
            manager.generated_records()[1].article,
            "16cut/Aurora Green AB/SS16/Non/203+",
        )
