"""Tests for the inventory splitting utility."""

from __future__ import annotations

from contextlib import redirect_stdout
from pathlib import Path
import io
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from tools import split_inventory


EXPECTED_SHEETS = list(split_inventory.SHEET_TO_FILE)


def _write_inventory_workbook(workbook_path: Path, missing_sheet: str | None = None) -> None:
    """Create a synthetic inventory workbook for splitter tests."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name in EXPECTED_SHEETS:
        if sheet_name == missing_sheet:
            continue

        sheet = workbook.create_sheet(sheet_name)
        sheet.merge_cells("A1:B1")
        sheet["A1"] = f"{sheet_name} header"
        sheet["A1"].font = Font(bold=True)
        sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        sheet["A2"] = "value"
        sheet["B2"] = 123

    unknown_sheet = workbook.create_sheet("в цапах")
    unknown_sheet["A1"] = "unknown"
    unknown_sheet["B1"] = 456

    workbook.save(workbook_path)
    workbook.close()


class TestSplitInventory(unittest.TestCase):
    """Coverage for splitting the consolidated inventory workbook."""

    def test_main_splits_inventory_into_input_files(self) -> None:
        """Verify known sheets are split, unknown sheets are reported, and formatting survives."""
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_file = temp_path / "Остатки на 04.07.2026.xlsx"
            _write_inventory_workbook(source_file, missing_sheet="12 граней холодные")

            input_dir = temp_path / "input"
            buffer = io.StringIO()
            with patch.object(split_inventory, "INPUT_DIR", input_dir), redirect_stdout(buffer):
                exit_code = split_inventory.main([str(source_file)])

            output = buffer.getvalue()

            self.assertEqual(exit_code, 0)
            self.assertIn("Unknown sheet:\nв цапах", output)
            self.assertIn("Skipped:\n12 граней холодные", output)
            self.assertIn("Inventory split completed", output)
            self.assertIn("Created:\n11", output)
            self.assertIn("Overwritten:\n0", output)
            self.assertIn("Skipped:\n1", output)
            self.assertIn("Unknown:\n1", output)

            split_file = input_dir / "K9_premium.xlsx"
            self.assertTrue(split_file.is_file())

            workbook = load_workbook(split_file)
            try:
                sheet = workbook.active
                self.assertEqual(workbook.sheetnames, ["K9 (премиум)"])
                self.assertIn("A1:B1", {str(range_ref) for range_ref in sheet.merged_cells.ranges})
                self.assertTrue(sheet["A1"].font.bold)
                self.assertEqual(sheet["A1"].fill.fgColor.rgb, "00FFF2CC")
                self.assertEqual(sheet["A2"].value, "value")
                self.assertEqual(sheet["B2"].value, 123)
            finally:
                workbook.close()

    def test_main_counts_existing_output_as_overwritten(self) -> None:
        """Verify existing output files are reported as overwritten, not created."""
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_file = temp_path / "Остатки sample.xlsx"

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "16 граней горячие"
            sheet["A1"] = "header"
            sheet["A2"] = "value"
            workbook.save(source_file)
            workbook.close()

            input_dir = temp_path / "input"
            input_dir.mkdir()
            existing_output = input_dir / "16cut_hot.xlsx"
            existing_output.write_bytes(b"old content")

            buffer = io.StringIO()
            with patch.object(split_inventory, "PROJECT_ROOT", temp_path), patch.object(
                split_inventory, "INPUT_DIR", input_dir
            ), redirect_stdout(buffer):
                exit_code = split_inventory.main([str(source_file)])

            output = buffer.getvalue()

            self.assertEqual(exit_code, 0)
            self.assertIn("Created:\n0", output)
            self.assertIn("Overwritten:\n1", output)
            self.assertNotEqual(existing_output.read_bytes(), b"old content")

    def test_main_discovers_inventory_file_in_project_root(self) -> None:
        """Verify the default CLI path discovery uses the project root inventory file."""
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_file = temp_path / "Остатки test.xlsx"
            _write_inventory_workbook(source_file, missing_sheet=None)

            input_dir = temp_path / "input"
            buffer = io.StringIO()
            with patch.object(split_inventory, "PROJECT_ROOT", temp_path), patch.object(
                split_inventory, "INPUT_DIR", input_dir
            ), redirect_stdout(buffer):
                exit_code = split_inventory.main([])

            output = buffer.getvalue()

            self.assertEqual(exit_code, 0)
            self.assertIn("Inventory split completed", output)
            self.assertIn("Created:\n12", output)
            self.assertTrue((input_dir / "16cut_hot.xlsx").is_file())
            self.assertTrue((input_dir / "K9.xlsx").is_file())
