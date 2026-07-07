"""Tests for NAS Parser Excel export."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
import tempfile
import unittest

from openpyxl import load_workbook

from nas_parser.domain import ProductRecord
from nas_parser.export import ExcelExporter


def make_record(
    *,
    sku: str | None,
    name: str | None,
    price: Decimal | None,
    quantity: Decimal | None,
    category: str | None,
    color: str,
    size: str,
    color_code: str | None = None,
    shape: str | None = None,
    fixation: str | None = None,
    cut: str | None = None,
) -> ProductRecord:
    """Create a ProductRecord fixture for exporter tests."""
    return ProductRecord(
        sku=sku,
        name=name,
        price=price,
        quantity=quantity,
        category=category,
        color=color,
        color_code=color_code,
        size=size,
        shape=shape,
        fixation=fixation,
        cut=cut,
        source_file=Path("input/source.xlsx"),
        source_sheet="Sheet1",
        source_row=1,
        parser_name="test",
    )


class TestExcelExporter(unittest.TestCase):
    """Coverage for ExcelExporter behavior."""

    def test_export_creates_workbook_with_products_sheet_and_rows(self) -> None:
        """Verify that exporter creates an xlsx file with headers and rows."""
        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "output.xlsx"
            exporter = ExcelExporter(output_file)
            records = [
                make_record(
                    sku="12cut/Crystal/SS16/Hot/001",
                    name="Stones 12 Crystal SS16 Hotfix",
                    price=Decimal("135"),
                    quantity=Decimal("374"),
                    category="Hotfix stones",
                    color="Crystal",
                    color_code="001",
                    size="SS16",
                    fixation="hot",
                    cut="12cut",
                ),
                make_record(
                    sku=None,
                    name=None,
                    price=Decimal("11"),
                    quantity=Decimal("1886"),
                    category=None,
                    color="CRYSTAL",
                    size="7*12",
                    shape="Drop",
                    fixation="sew",
                    cut=None,
                ),
            ]

            result = exporter.export(records)

            self.assertEqual(result, output_file)
            self.assertTrue(output_file.is_file())

            workbook = load_workbook(output_file, read_only=True, data_only=True)
            try:
                self.assertIn("Товары", workbook.sheetnames)
                sheet = workbook["Товары"]
                rows = list(sheet.iter_rows(values_only=True))
            finally:
                workbook.close()

        self.assertEqual(rows[0], ExcelExporter.HEADER)
        self.assertEqual(len(rows), 3)
        self.assertEqual(
            rows[1],
            (
                None,
                None,
                11,
                1886,
                None,
                "CRYSTAL",
                "7*12",
                "Drop",
                "K9",
                None,
            ),
        )
        self.assertEqual(
            rows[2],
            (
                "12cut/Crystal/SS16/Hot/001",
                "Stones 12 Crystal SS16 Hotfix",
                135,
                374,
                "Hotfix stones",
                "001",
                "SS16",
                None,
                "Hot",
                "12cut",
            ),
        )

    def test_export_falls_back_to_color_name_without_color_code(self) -> None:
        """Verify that color names are exported when no color code exists."""
        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "fallback.xlsx"
            records = [
                make_record(
                    sku="Aquamarine/Rivoli/12x12",
                    name="Пришивные стразы Aquamarine Rivoli 12*12",
                    price=Decimal("20"),
                    quantity=Decimal("10"),
                    category="Пришивные стразы",
                    color="Aquamarine",
                    size="12*12",
                    shape="Rivoli",
                    fixation="sew",
                ),
            ]

            ExcelExporter(output_file).export(records)

            workbook = load_workbook(output_file, read_only=True, data_only=True)
            try:
                row = list(workbook["Товары"].iter_rows(values_only=True))[1]
            finally:
                workbook.close()

        self.assertEqual(row[5], "Aquamarine")
        self.assertEqual(row[8], "K9")

    def test_export_sorts_records_deterministically(self) -> None:
        """Verify that records are sorted by stable ProductRecord fields."""
        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "sorted.xlsx"
            records = [
                make_record(
                    sku="sku-3",
                    name="Name 3",
                    price=Decimal("3"),
                    quantity=Decimal("3"),
                    category="B",
                    color="Blue",
                    size="SS8",
                ),
                make_record(
                    sku="sku-2",
                    name="Name 2",
                    price=Decimal("2"),
                    quantity=Decimal("2"),
                    category="A",
                    color="Topaz",
                    size="SS6",
                ),
                make_record(
                    sku="sku-1",
                    name="Name 1",
                    price=Decimal("1"),
                    quantity=Decimal("1"),
                    category="A",
                    color="Crystal",
                    size="SS16",
                ),
            ]

            ExcelExporter(output_file).export(records)

            workbook = load_workbook(output_file, read_only=True, data_only=True)
            try:
                sheet = workbook["Товары"]
                rows = list(sheet.iter_rows(values_only=True))
            finally:
                workbook.close()

        self.assertEqual([row[0] for row in rows[1:]], ["sku-1", "sku-2", "sku-3"])

    def test_export_rounds_prices_up_to_whole_rubles(self) -> None:
        """Verify that exported prices are always rounded upward to whole rubles."""
        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = Path(temp_dir) / "prices.xlsx"
            records = [
                make_record(
                    sku="sku-1800001",
                    name="Price 18.00001",
                    price=Decimal("18.00001"),
                    quantity=Decimal("1"),
                    category="A",
                    color="A",
                    size="1",
                ),
                make_record(
                    sku="sku-185",
                    name="Price 18.5",
                    price=Decimal("18.5"),
                    quantity=Decimal("1"),
                    category="A",
                    color="B",
                    size="1",
                ),
                make_record(
                    sku="sku-18999",
                    name="Price 18.999",
                    price=Decimal("18.999"),
                    quantity=Decimal("1"),
                    category="A",
                    color="C",
                    size="1",
                ),
                make_record(
                    sku="sku-19",
                    name="Price 19",
                    price=Decimal("19"),
                    quantity=Decimal("1"),
                    category="A",
                    color="D",
                    size="1",
                ),
                make_record(
                    sku="sku-0",
                    name="Price 0",
                    price=Decimal("0"),
                    quantity=Decimal("1"),
                    category="A",
                    color="E",
                    size="1",
                ),
                make_record(
                    sku="sku-01",
                    name="Price 0.1",
                    price=Decimal("0.1"),
                    quantity=Decimal("1"),
                    category="A",
                    color="F",
                    size="1",
                ),
                make_record(
                    sku="sku-none",
                    name="Price None",
                    price=None,
                    quantity=Decimal("1"),
                    category="A",
                    color="G",
                    size="1",
                ),
            ]

            ExcelExporter(output_file).export(records)

            workbook = load_workbook(output_file, read_only=True, data_only=True)
            try:
                rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))
            finally:
                workbook.close()

        exported_prices = {row[0]: row[2] for row in rows[1:]}
        self.assertEqual(exported_prices["sku-1800001"], 19)
        self.assertEqual(exported_prices["sku-185"], 19)
        self.assertEqual(exported_prices["sku-18999"], 19)
        self.assertEqual(exported_prices["sku-19"], 19)
        self.assertEqual(exported_prices["sku-0"], 0)
        self.assertEqual(exported_prices["sku-01"], 1)
        self.assertIsNone(exported_prices["sku-none"])
        self.assertEqual(records[0].price, Decimal("18.00001"))
