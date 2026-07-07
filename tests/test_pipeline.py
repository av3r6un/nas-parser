"""Tests for the integrated NAS Parser pipeline."""

from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from nas_parser.config import AppConfig
from nas_parser.pipeline import Pipeline


def _write_color_reference_workbook(workbook_path: Path) -> None:
    """Create a small color reference workbook for pipeline tests."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(
        (
            "\u0410\u0440\u0442\u0438\u043a\u0443\u043b WB \u0438 Ozon",
            "\u0413\u0440\u0430\u043d\u0438",
            "\u0426\u0432\u0435\u0442",
            "\u0420\u0430\u0437\u043c\u0435\u0440",
            "\u0422\u0438\u043f",
            "\u0410\u0440\u0442\u0438\u043a\u0443\u043b \u0446\u0432\u0435\u0442\u0430",
        )
    )
    sheet.append(("12cut/Crystal/SS3/Hot/001", "12cut", "Crystal", "SS3", "Hot", "001"))
    sheet.append(
        ("16cut/Crystal AB/SS30/Non/001+", "16cut", "Crystal AB", "SS30", "Non", "001+")
    )
    sheet.append(("16cut/New AB/SS16/Hot/112", "16cut", "New AB", "SS16", "Hot", "112"))
    workbook.save(workbook_path)
    workbook.close()


def _write_cut_workbook(
    workbook_path: Path,
    *,
    sheet_title: str,
    title_row: tuple[object, ...],
    color_row: tuple[object, ...],
    second_row: tuple[object, ...],
) -> None:
    """Create a small cut workbook for pipeline tests."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.append(title_row)
    sheet.append(("color", "size", "pack_size", "price", "received", "sold", "stock", "color_code"))
    sheet.append(color_row)
    sheet.append(second_row)
    workbook.save(workbook_path)
    workbook.close()


class TestPipeline(unittest.TestCase):
    """Coverage for the first working pipeline."""

    def test_pipeline_processes_supported_files_and_skips_unsupported_ones(self) -> None:
        """Verify the pipeline runs end to end across supported and unsupported files."""
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir) / "input"
            reference_dir = Path(temp_dir) / "reference"
            output_dir = Path(temp_dir) / "output"
            output_file = output_dir / "catalog.xlsx"
            logs_dir = Path(temp_dir) / "logs"
            input_dir.mkdir()
            reference_dir.mkdir()
            output_dir.mkdir()

            _write_color_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            _write_cut_workbook(
                input_dir / "12cut_hot.xlsx",
                sheet_title="12cut_hot",
                title_row=("hot fixation", None, None, None, None, None, None, None),
                color_row=("Crystal", "SS3", 1440, 135, 516, 142, 374, "001"),
                second_row=(None, "SS4", 1440, 135, 600, 63, 537, None),
            )
            _write_cut_workbook(
                input_dir / "16cut_non.xlsx",
                sheet_title="16cut_non",
                title_row=("non fixation", None, None, None, None, None, None, None),
                color_row=("CRYSTAL", "SS6", 1440, 590, 57, 1, 56, "001"),
                second_row=(None, "SS8", 1440, 701, 65, 2, 63, None),
            )

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "K9"
            sheet.append(("K9 title", None, None, None, None, None, None, None))
            sheet.append(("color", "shape", "size", "price", "pack", "received", "sold", "stock"))
            sheet.append(("CRYSTAL", "Drop", "7*12", 11, 72, 1886, None, 1886))
            workbook.save(input_dir / "K9.xlsx")
            workbook.close()

            pipeline = Pipeline(
                AppConfig(
                    input_dir=input_dir,
                    reference_dir=reference_dir,
                    output_dir=output_dir,
                    output_file=output_file,
                    logs_dir=logs_dir,
                )
            )
            records, report = pipeline.run()

            self.assertTrue(output_file.is_file())
            workbook = load_workbook(output_file, read_only=True, data_only=True)
            try:
                self.assertIn("\u0422\u043e\u0432\u0430\u0440\u044b", workbook.sheetnames)
                sheet = workbook["\u0422\u043e\u0432\u0430\u0440\u044b"]
                output_rows = list(sheet.iter_rows(values_only=True))
            finally:
                workbook.close()

        self.assertEqual(len(records), 5)
        self.assertGreater(len(records), 0)
        self.assertGreater(len(output_rows), 1)
        self.assertEqual(records[0].color_code, "001")
        self.assertEqual(records[0].sku, "12cut/Crystal/SS3/Hot/001")
        self.assertEqual(records[0].category, "Стразы горячей фиксации")
        self.assertEqual(records[0].name, "Crystal")
        self.assertEqual(records[2].color_code, "001")
        self.assertEqual(records[2].sku, "16cut/Crystal/SS6/Non/001")
        self.assertEqual(records[2].category, "Стразы холодной фиксации")
        self.assertEqual(records[2].name, "Crystal")
        self.assertEqual(records[4].parser_name, "k9")
        self.assertEqual(records[4].color, "CRYSTAL")
        self.assertEqual(records[4].shape, "Drop")
        self.assertEqual(records[4].size, "7x12")
        self.assertEqual(records[4].fixation, "sew")
        self.assertIsNone(records[4].cut)
        self.assertEqual(records[4].sku, "K9/Crystal/Drop/7x12/001")
        self.assertEqual(records[4].name, "Crystal Drop")
        self.assertEqual(records[4].category, "Пришивные стразы")
        self.assertEqual(records[4].color_code, "001")
        self.assertIn("files_found=3", report.summary())
        self.assertIn("files_processed=3", report.summary())
        self.assertIn("files_skipped=0", report.summary())
        self.assertIn("records_created=5", report.summary())
        self.assertIn("warnings=0", report.summary())
        self.assertIn("errors=0", report.summary())
        self.assertIn(f"output_file={output_file}", report.summary())

    def test_pipeline_generates_reference_file_for_missing_color(self) -> None:
        """Verify missing K9 colors are generated and written to a new reference file."""
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir) / "input"
            reference_dir = Path(temp_dir) / "reference"
            output_dir = Path(temp_dir) / "output"
            output_file = output_dir / "catalog.xlsx"
            logs_dir = Path(temp_dir) / "logs"
            input_dir.mkdir()
            reference_dir.mkdir()
            output_dir.mkdir()

            base_reference = reference_dir / "colorcode-articul.xlsx"
            _write_color_reference_workbook(base_reference)
            base_bytes = base_reference.read_bytes()

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "K9"
            sheet.append(("K9 title", None, None, None, None, None, None, None))
            sheet.append(("color", "shape", "size", "price", "pack", "received", "sold", "stock"))
            sheet.append(("AURORA GREEN", "Navette", "7*15", 11, 72, 1886, None, 1886))
            workbook.save(input_dir / "K9.xlsx")
            workbook.close()

            pipeline = Pipeline(
                AppConfig(
                    input_dir=input_dir,
                    reference_dir=reference_dir,
                    output_dir=output_dir,
                    output_file=output_file,
                    logs_dir=logs_dir,
                )
            )
            records, report = pipeline.run()

            generated_file = reference_dir / "generated" / "colorcode-articul_gen1.xlsx"
            generated_workbook = load_workbook(generated_file, read_only=True, data_only=True)
            try:
                rows = list(generated_workbook.active.iter_rows(values_only=True))
            finally:
                generated_workbook.close()
            generated_exists = generated_file.is_file()
            base_after_bytes = base_reference.read_bytes()

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].color_code, "203")
        self.assertEqual(records[0].sku, "K9/Aurora Green/Navette/7x15/203")
        self.assertTrue(generated_exists)
        self.assertEqual(base_after_bytes, base_bytes)
        self.assertIn(("K9/Aurora Green/Navette/7x15/203", "K9", "Aurora Green", "7x15", "K9", "203"), rows)
        self.assertIn("Reference updates:", report.logs())
        self.assertIn("Generated colors: 1", report.logs())
        self.assertIn(f"Generated reference: {generated_file}", report.logs())
        self.assertIn("warnings=0", report.summary())

    def test_pipeline_uses_latest_generated_reference_file(self) -> None:
        """Verify the active color reference comes from the latest generation file."""
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir) / "input"
            reference_dir = Path(temp_dir) / "reference"
            generated_dir = reference_dir / "generated"
            output_dir = Path(temp_dir) / "output"
            output_file = output_dir / "catalog.xlsx"
            logs_dir = Path(temp_dir) / "logs"
            input_dir.mkdir()
            reference_dir.mkdir()
            generated_dir.mkdir()
            output_dir.mkdir()

            _write_color_reference_workbook(reference_dir / "colorcode-articul.xlsx")
            gen5 = generated_dir / "colorcode-articul_gen5.xlsx"
            _write_color_reference_workbook(gen5)
            workbook = load_workbook(gen5)
            try:
                workbook.active.append(("K9/Crystal/Drop/7x12/777", "K9", "Crystal", "7x12", "K9", "777"))
                workbook.save(gen5)
            finally:
                workbook.close()

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "K9"
            sheet.append(("K9 title", None, None, None, None, None, None, None))
            sheet.append(("color", "shape", "size", "price", "pack", "received", "sold", "stock"))
            sheet.append(("Crystal", "Drop", "7*12", 11, 72, 1886, None, 1886))
            workbook.save(input_dir / "K9.xlsx")
            workbook.close()

            pipeline = Pipeline(
                AppConfig(
                    input_dir=input_dir,
                    reference_dir=reference_dir,
                    output_dir=output_dir,
                    output_file=output_file,
                    logs_dir=logs_dir,
                )
            )
            records, report = pipeline.run()

        self.assertEqual(records[0].color_code, "777")
        self.assertEqual(records[0].sku, "K9/Crystal/Drop/7x12/777")
        self.assertFalse((generated_dir / "colorcode-articul_gen6.xlsx").exists())
        self.assertIn("Reference updates:", report.logs())
        self.assertIn("No changes", report.logs())

    def test_pipeline_repeated_run_does_not_create_new_generation(self) -> None:
        """Verify repeated runs on identical input do not create new reference generations."""
        with tempfile.TemporaryDirectory() as temp_dir:
            input_dir = Path(temp_dir) / "input"
            reference_dir = Path(temp_dir) / "reference"
            generated_dir = reference_dir / "generated"
            output_dir = Path(temp_dir) / "output"
            output_file = output_dir / "catalog.xlsx"
            logs_dir = Path(temp_dir) / "logs"
            input_dir.mkdir()
            reference_dir.mkdir()
            output_dir.mkdir()

            _write_color_reference_workbook(reference_dir / "colorcode-articul.xlsx")

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "K9"
            sheet.append(("K9 title", None, None, None, None, None, None, None))
            sheet.append(("color", "shape", "size", "price", "pack", "received", "sold", "stock"))
            sheet.append(("AURORA GREEN", "Navette", "7*15", 11, 72, 1886, None, 1886))
            workbook.save(input_dir / "K9.xlsx")
            workbook.close()

            pipeline = Pipeline(
                AppConfig(
                    input_dir=input_dir,
                    reference_dir=reference_dir,
                    output_dir=output_dir,
                    output_file=output_file,
                    logs_dir=logs_dir,
                )
            )
            first_records, first_report = pipeline.run()
            second_records, second_report = pipeline.run()

            gen1 = generated_dir / "colorcode-articul_gen1.xlsx"
            gen2 = generated_dir / "colorcode-articul_gen2.xlsx"
            gen1_exists = gen1.is_file()
            gen2_exists = gen2.exists()
            workbook = load_workbook(gen1, read_only=True, data_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()

        self.assertEqual(first_records[0].color_code, "203")
        self.assertEqual(second_records[0].color_code, "203")
        self.assertTrue(gen1_exists)
        self.assertFalse(gen2_exists)
        self.assertEqual(len([row for row in rows if row[2] == "Aurora Green"]), 1)
        self.assertIn("Generated colors: 1", first_report.logs())
        self.assertIn("No changes", second_report.logs())
