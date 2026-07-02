"""Excel reader implementation for NAS Parser."""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from openpyxl import load_workbook

from nas_parser.readers.base import ReaderBase
from nas_parser.source import SourceRow


class ExcelReader(ReaderBase):
    """Read rows from Excel workbooks and convert them into SourceRow objects."""

    def __init__(self, source_file: Path) -> None:
        """Initialize the reader for a specific Excel workbook path."""
        self._source_file = source_file

    @property
    def name(self) -> str:
        """Return the canonical reader name."""
        return "excel"

    def read(self) -> Iterable[SourceRow]:
        """Yield source rows from all sheets in the workbook."""
        workbook = load_workbook(self._source_file, read_only=True, data_only=False)

        try:
            for sheet in workbook.worksheets:
                for row_index, row in enumerate(sheet.iter_rows(), start=1):
                    values = tuple(cell.value for cell in row)
                    if all(value is None for value in values):
                        continue

                    yield SourceRow(
                        values=values,
                        source_file=self._source_file,
                        source_sheet=sheet.title,
                        source_row=row_index,
                    )
        finally:
            workbook.close()
