"""Color reference file management utilities."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import shutil
import unicodedata

from openpyxl import load_workbook


@dataclass(slots=True)
class GeneratedColorRecord:
    """Represent a color reference row generated in memory."""

    article: str
    cut: str | None
    color: str
    size: str | None
    type: str
    color_code: str


class ColorReferenceManager:
    """Manage color reference workbook files and generated copies."""

    BASE_FILE_NAME = "colorcode-articul.xlsx"
    GENERATED_DIR_NAME = "generated"
    GENERATED_FILE_PATTERN = re.compile(r"^colorcode-articul_gen(\d+)\.xlsx$")
    START_GENERATED_CODE = 203
    COLOR_HEADERS = {"Цвет", "Р¦РІРµС‚"}
    CODE_HEADERS = {"Артикул цвета", "РђСЂС‚РёРєСѓР» С†РІРµС‚Р°"}
    ARTICLE_HEADERS = {"Артикул WB/Ozon", "Артикул WB и Ozon"}
    CUT_HEADERS = {"Грани"}
    SIZE_HEADERS = {"Размер"}
    TYPE_HEADERS = {"Тип"}

    def __init__(self, reference_dir: Path) -> None:
        """Initialize the manager with a reference directory."""
        self._reference_dir = reference_dir
        self._generated_records: list[GeneratedColorRecord] = []
        self._existing_color_codes: dict[str, str] | None = None

    def get_active_reference_file(self) -> Path:
        """Return the latest generated reference file or the base reference file."""
        generation_files = self._generation_files()
        if not generation_files:
            return self._base_reference_file()

        return max(generation_files, key=self._generation_number)

    def next_generation_file(self) -> Path:
        """Return the next generated reference file path, creating the directory."""
        generated_dir = self._generated_dir()
        generated_dir.mkdir(parents=True, exist_ok=True)

        generation_files = self._generation_files()
        next_number = 1
        if generation_files:
            next_number = max(self._generation_number(path) for path in generation_files) + 1

        return generated_dir / f"colorcode-articul_gen{next_number}.xlsx"

    def copy_reference_for_generation(self) -> Path:
        """Copy the active reference workbook into the next generated file."""
        source_file = self.get_active_reference_file()
        destination_file = self.next_generation_file()
        shutil.copy2(source_file, destination_file)
        return destination_file

    def write_generated_records(self, generation_file: Path) -> Path:
        """Append generated color records to a generation workbook."""
        if not self._generated_records:
            return generation_file

        workbook = load_workbook(generation_file)
        try:
            sheet = workbook.active
            headers = self._read_headers(sheet.iter_rows(values_only=True))
            if headers is None:
                return generation_file

            indexes = self._reference_column_indexes(headers)
            existing_colors = self._existing_normalized_colors_in_sheet(sheet, indexes["color"])
            records_to_write: list[GeneratedColorRecord] = []
            for record in self._generated_records:
                normalized_color = self._normalize_color(record.color)
                if normalized_color in existing_colors:
                    continue

                records_to_write.append(record)
                existing_colors.add(normalized_color)

            if not records_to_write:
                return generation_file

            if (
                self._generation_number(generation_file) == 1
                and not self._has_generated_records_in_sheet(sheet, indexes["color_code"])
            ):
                row_width = max(indexes.values()) + 1
                for _ in range(100):
                    sheet.append([None] * row_width)

            for record in records_to_write:
                row = [None] * (max(indexes.values()) + 1)
                row[indexes["article"]] = record.article
                row[indexes["cut"]] = "K9" if record.type == "K9" else record.cut
                row[indexes["color"]] = record.color
                row[indexes["size"]] = record.size
                row[indexes["type"]] = record.type
                row[indexes["color_code"]] = self._format_color_code_display(record.color_code)
                sheet.append(row)

            workbook.save(generation_file)
        finally:
            workbook.close()

        return generation_file

    def ensure_color(
        self,
        color: str,
        cut: str | None,
        size: str | None,
        fixation: str,
    ) -> str:
        """Return an existing color code or generate a new one in memory."""
        normalized_color = self._normalize_color(color)

        existing_code = self._existing_color_codes_by_color().get(normalized_color)
        if existing_code is not None:
            return existing_code

        generated_code = self._generated_code_for_color(normalized_color)
        if generated_code is not None:
            return generated_code

        color_code = self._new_color_code(color)
        display_color = self._canonical_color_display(color)
        self._generated_records.append(
            GeneratedColorRecord(
                article=self._build_article(
                    color=display_color,
                    cut=cut,
                    size=size,
                    fixation=fixation,
                    color_code=color_code,
                ),
                cut=self._format_shape_display(cut) if fixation == "sew" else (cut.strip() if isinstance(cut, str) else cut),
                color=display_color,
                size=size.strip() if isinstance(size, str) else size,
                type=self._format_type(fixation),
                color_code=color_code,
            )
        )
        return color_code

    def generated_records(self) -> tuple[GeneratedColorRecord, ...]:
        """Return color reference records generated during the current run."""
        return tuple(self._generated_records)

    def _base_reference_file(self) -> Path:
        """Return the base color reference workbook path."""
        return self._reference_dir / self.BASE_FILE_NAME

    def _generated_dir(self) -> Path:
        """Return the generated references directory path."""
        return self._reference_dir / self.GENERATED_DIR_NAME

    def _generation_files(self) -> list[Path]:
        """Return generated reference files matching the expected naming pattern."""
        generated_dir = self._generated_dir()
        if not generated_dir.is_dir():
            return []

        return [
            path
            for path in generated_dir.iterdir()
            if path.is_file() and self.GENERATED_FILE_PATTERN.fullmatch(path.name)
        ]

    @classmethod
    def _generation_number(cls, path: Path) -> int:
        """Return the generation number encoded in a generated reference filename."""
        match = cls.GENERATED_FILE_PATTERN.fullmatch(path.name)
        return int(match.group(1)) if match is not None else 0

    def _existing_color_codes_by_color(self) -> dict[str, str]:
        """Return color codes loaded from the active reference workbook."""
        if self._existing_color_codes is None:
            self._existing_color_codes = self._load_existing_color_codes()

        return self._existing_color_codes

    def _load_existing_color_codes(self) -> dict[str, str]:
        """Load existing color codes from the active reference workbook."""
        source_file = self.get_active_reference_file()
        if not source_file.is_file():
            return {}

        workbook = load_workbook(source_file, read_only=True, data_only=True)
        try:
            color_codes: dict[str, str] = {}
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                headers = self._read_headers(rows)
                if headers is None:
                    continue

                color_index = self._header_index(headers, self.COLOR_HEADERS)
                code_index = self._header_index(headers, self.CODE_HEADERS)
                if color_index is None or code_index is None:
                    continue

                for row in rows:
                    color_value = self._row_value(row, color_index)
                    code_value = self._row_value(row, code_index)
                    if color_value is None or code_value is None:
                        continue

                    color_text = self._canonical_color_display(str(color_value))
                    code_text = str(code_value).strip()
                    if color_text and code_text:
                        color_codes[self._normalize_color(color_text)] = code_text

            return color_codes
        finally:
            workbook.close()

    def _new_color_code(self, color: str) -> str:
        """Return the next generated color code for a missing color."""
        base_color = self._base_ab_color(color)
        if base_color is not None:
            base_code = self._code_for_any_color(base_color)
            if base_code is not None:
                return f"{self._numeric_code(base_code)}+"

        next_number = max([self.START_GENERATED_CODE - 1, *self._used_generated_numbers()]) + 1
        return str(next_number) + ("+" if base_color is not None else "")

    def _code_for_any_color(self, color: str) -> str | None:
        """Return an existing or generated code for a color."""
        normalized_color = self._normalize_color(color)
        existing_code = self._existing_color_codes_by_color().get(normalized_color)
        if existing_code is not None:
            return existing_code

        return self._generated_code_for_color(normalized_color)

    def _generated_code_for_color(self, normalized_color: str) -> str | None:
        """Return an in-memory generated code for a normalized color."""
        for record in self._generated_records:
            if self._normalize_color(record.color) == normalized_color:
                return record.color_code

        return None

    def _used_generated_numbers(self) -> list[int]:
        """Return numeric generated code values from existing and in-memory records."""
        codes = list(self._existing_color_codes_by_color().values())
        codes.extend(record.color_code for record in self._generated_records)
        return [
            number
            for number in (self._numeric_code(code) for code in codes)
            if number >= self.START_GENERATED_CODE
        ]

    @staticmethod
    def _read_headers(rows: object) -> dict[str, int] | None:
        """Read a header row and map header names to column indexes."""
        for row in rows:
            if all(value is None for value in row):
                continue

            return {
                str(value).strip(): index
                for index, value in enumerate(row)
                if value is not None and str(value).strip()
            }

        return None

    @staticmethod
    def _header_index(headers: dict[str, int], names: set[str]) -> int | None:
        """Return a header index by any accepted header name."""
        for name in names:
            if name in headers:
                return headers[name]

        return None

    def _reference_column_indexes(self, headers: dict[str, int]) -> dict[str, int]:
        """Return output column indexes for generated reference records."""
        indexes = {
            "article": self._header_index(headers, self.ARTICLE_HEADERS),
            "cut": self._header_index(headers, self.CUT_HEADERS),
            "color": self._header_index(headers, self.COLOR_HEADERS),
            "size": self._header_index(headers, self.SIZE_HEADERS),
            "type": self._header_index(headers, self.TYPE_HEADERS),
            "color_code": self._header_index(headers, self.CODE_HEADERS),
        }
        if any(index is None for index in indexes.values()):
            missing = ", ".join(name for name, index in indexes.items() if index is None)
            raise ValueError(f"Missing reference columns: {missing}")

        return {name: index for name, index in indexes.items() if index is not None}

    def _existing_normalized_colors_in_sheet(self, sheet: object, color_index: int) -> set[str]:
        """Return normalized color names already present in a reference sheet."""
        colors: set[str] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            color_value = self._row_value(row, color_index)
            if color_value is None:
                continue

            color_text = str(color_value).strip()
            if color_text:
                colors.add(self._normalize_color(color_text))

        return colors

    def _has_generated_records_in_sheet(self, sheet: object, color_code_index: int) -> bool:
        """Return whether the sheet already contains generated color code rows."""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            code_value = self._row_value(row, color_code_index)
            if code_value is None:
                continue

            if self._numeric_code(str(code_value)) >= self.START_GENERATED_CODE:
                return True

        return False

    @staticmethod
    def _row_value(row: tuple[object, ...], index: int) -> object | None:
        """Return a row value when the column exists."""
        return row[index] if index < len(row) else None

    @staticmethod
    def _normalize_color(color: str) -> str:
        """Normalize color names for manager lookups."""
        return ColorReferenceManager._canonical_color_display(color).casefold()

    @staticmethod
    def _canonical_color_display(color: str) -> str:
        """Return a display color with normalized Unicode and whitespace."""
        normalized = unicodedata.normalize("NFKC", color)
        return " ".join(normalized.split())

    @staticmethod
    def _format_shape_display(shape: str | None) -> str | None:
        """Return a normalized title-cased shape display value."""
        if shape is None:
            return None

        normalized = unicodedata.normalize("NFKC", shape)
        parts = normalized.split()
        if not parts:
            return ""

        return " ".join(part[:1].upper() + part[1:].lower() for part in parts)

    @staticmethod
    def _base_ab_color(color: str) -> str | None:
        """Return the base color when the color is an AB variant."""
        text = ColorReferenceManager._canonical_color_display(color)
        return text[:-3] if text.casefold().endswith(" ab") else None

    @staticmethod
    def _numeric_code(code: str) -> int:
        """Return the numeric part of a color code."""
        text = code.strip().removesuffix("+")
        return int(text) if text.isdecimal() else 0

    @staticmethod
    def _format_color_code_display(color_code: str) -> str:
        """Return a display color code with zero padding below 100."""
        text = color_code.strip()
        suffix = "+" if text.endswith("+") else ""
        core = text[:-1] if suffix else text
        if core.isdecimal():
            return core.zfill(3) + suffix

        return text

    @staticmethod
    def _format_type(fixation: str) -> str:
        """Return the reference type value for a fixation."""
        if fixation == "hot":
            return "Hot"
        if fixation == "non":
            return "Non"
        if fixation == "sew":
            return "K9"

        return fixation

    def _build_article(
        self,
        *,
        color: str,
        cut: str | None,
        size: str | None,
        fixation: str,
        color_code: str,
    ) -> str:
        """Build the reference article for a generated color row."""
        color_text = self._canonical_color_display(color)
        size_text = size.strip() if isinstance(size, str) else ""
        code_text = self._format_color_code_display(color_code)

        if fixation == "sew":
            shape_text = self._format_shape_display(cut) if isinstance(cut, str) else ""
            return "/".join(("K9", color_text, shape_text or "", size_text, code_text))

        cut_text = cut.strip() if isinstance(cut, str) else ""
        return "/".join((cut_text, color_text, size_text, self._format_type(fixation), code_text))
