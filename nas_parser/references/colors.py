"""Color reference contract for NAS Parser."""

from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass, field
from openpyxl import load_workbook
from pathlib import Path


@dataclass(slots=True)
class ColorReference:
    """Represent loaded color codes and provide fast lookup by color name."""

    source_file: Path | None = field(default=None)
    loaded: bool = field(default=False, init=False)
    _color_to_code: dict[str, str] = field(default_factory=dict, repr=False)

    @property
    def name(self) -> str:
        """Return the canonical name of the color reference."""
        return "colors"

    def get_code(self, color_name: str) -> str | None:
        """Return the color code for a color name, if it exists."""
        return self._color_to_code.get(self._normalize_key(color_name))

    @staticmethod
    def _normalize_key(value: str) -> str:
        """Normalize a lookup key just enough for stable reference matching."""
        return value.strip().casefold()


class ColorReferenceLoader:
    """Load the color reference from an Excel workbook."""

    def __init__(self, source_file: Path) -> None:
        """Initialize the loader for the color reference workbook."""
        self._source_file = source_file

    def load(self) -> ColorReference:
        """Load the color reference from the configured workbook path."""
        workbook = load_workbook(self._source_file, read_only=True, data_only=True)

        try:
            color_to_code: dict[str, str] = {}

            for sheet in workbook.worksheets:
                headers = self._read_headers(sheet.iter_rows(values_only=True))
                if headers is None:
                    continue

                color_index = headers.get("Цвет")
                code_index = headers.get("Артикул цвета")

                if color_index is None or code_index is None:
                    continue

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if self._is_empty_row(row):
                        continue

                    color_value = row[color_index]
                    code_value = row[code_index]
                    if color_value is None or code_value is None:
                        continue

                    color_name = str(color_value).strip()
                    color_code = str(code_value).strip()
                    if not color_name or not color_code:
                        continue

                    color_to_code[color_name.casefold()] = color_code

            reference = ColorReference(
                source_file=self._source_file,
                _color_to_code=color_to_code,
            )
            reference.loaded = True
            return reference
        finally:
            workbook.close()

    @staticmethod
    def _read_headers(
        rows: Iterable[tuple[object, ...]]
    ) -> dict[str, int] | None:
        """Read a header row and map header names to zero-based column indexes."""
        for row in rows:
            if ColorReferenceLoader._is_empty_row(row):
                continue

            return {
                str(value).strip(): index
                for index, value in enumerate(row)
                if value is not None and str(value).strip()
            }

        return None

    @staticmethod
    def _is_empty_row(row: tuple[object, ...]) -> bool:
        """Return True when the row contains no meaningful values."""
        return all(cell is None for cell in row)
