"""Catalog audit utilities for exported NAS Parser workbooks."""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
import sys

from openpyxl import load_workbook


SKU_COLUMN = 0
NAME_COLUMN = 1
PRICE_COLUMN = 2
QUANTITY_COLUMN = 3
CATEGORY_COLUMN = 4
SHAPE_COLUMN = 7
FIXATION_COLUMN = 8
CUT_COLUMN = 9


@dataclass(slots=True)
class AuditReport:
    """Collected statistics about an exported catalog workbook."""

    rows: int = 0
    empty_sku: int = 0
    duplicate_sku: int = 0
    empty_name: int = 0
    category_counts: dict[str, int] = field(default_factory=dict)
    empty_price: int = 0
    negative_price: int = 0
    empty_quantity: int = 0
    negative_quantity: int = 0
    formula_quantity: int = 0
    empty_shape: int = 0
    empty_fixation: int = 0
    empty_cut: int = 0
    duplicate_sku_rows: dict[str, list[int]] = field(default_factory=dict)
    empty_sku_rows: list[int] = field(default_factory=list)
    empty_price_rows: list[int] = field(default_factory=list)
    empty_quantity_rows: list[int] = field(default_factory=list)

    def summary(self) -> str:
        """Return a readable multi-line summary of the audit result."""
        lines = [
            f"Rows: {self.rows}",
            "",
            "SKU:",
            f"  empty: {self.empty_sku}",
            f"  duplicates: {self.duplicate_sku}",
            "",
            "Name:",
            f"  empty: {self.empty_name}",
            "",
            "Category:",
        ]

        if self.category_counts:
            lines.extend(
                f"  {category}: {count}"
                for category, count in sorted(self.category_counts.items())
            )
        else:
            lines.append("  none: 0")

        lines.extend(
            (
                "",
                "Price:",
                f"  empty: {self.empty_price}",
                f"  negative: {self.negative_price}",
                "",
                "Quantity:",
                f"  empty: {self.empty_quantity}",
                f"  negative: {self.negative_quantity}",
                f"  formulas: {self.formula_quantity}",
                "",
                "Shape:",
                f"  empty: {self.empty_shape}",
                "",
                "Fixation:",
                f"  empty: {self.empty_fixation}",
                "",
                "Cut:",
                f"  empty: {self.empty_cut}",
            )
        )

        detail_lines = self._detail_lines()
        if detail_lines:
            lines.extend(("", *detail_lines))

        return "\n".join(lines)

    def _detail_lines(self) -> list[str]:
        """Return detailed row-level diagnostics for detected problems."""
        lines: list[str] = []

        if self.duplicate_sku_rows:
            lines.append("Duplicate SKU:")
            for sku, rows in sorted(self.duplicate_sku_rows.items()):
                lines.extend(
                    (
                        sku,
                        "Rows:",
                        *(str(row) for row in rows),
                        "",
                    )
                )
            lines.pop()

        if self.empty_sku_rows:
            if lines:
                lines.append("")
            lines.extend(
                (
                    "Empty SKU:",
                    "Rows:",
                    *(str(row) for row in self.empty_sku_rows),
                )
            )

        if self.empty_price_rows:
            if lines:
                lines.append("")
            lines.extend(
                (
                    "Empty Price:",
                    "Rows:",
                    *(str(row) for row in self.empty_price_rows),
                )
            )

        if self.empty_quantity_rows:
            if lines:
                lines.append("")
            lines.extend(
                (
                    "Empty Quantity:",
                    "Rows:",
                    *(str(row) for row in self.empty_quantity_rows),
                )
            )

        return lines


class CatalogAudit:
    """Audit an exported catalog workbook without changing its contents."""

    def __init__(self, excel_file: Path) -> None:
        """Store the exported workbook path."""
        self._excel_file = excel_file

    def run(self) -> AuditReport:
        """Collect quality statistics from the exported workbook."""
        report = AuditReport()
        duplicate_candidates: dict[str, list[int]] = {}

        workbook = load_workbook(self._excel_file, read_only=True, data_only=False)
        try:
            sheet = workbook.active
            for excel_row, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=False),
                start=2,
            ):
                if self._is_empty_row(row):
                    continue

                report.rows += 1

                sku = self._string_value(self._cell_value(row, SKU_COLUMN))
                name = self._string_value(self._cell_value(row, NAME_COLUMN))
                category = self._string_value(self._cell_value(row, CATEGORY_COLUMN))
                shape = self._string_value(self._cell_value(row, SHAPE_COLUMN))
                fixation = self._string_value(self._cell_value(row, FIXATION_COLUMN))
                cut = self._string_value(self._cell_value(row, CUT_COLUMN))

                if sku is None:
                    report.empty_sku += 1
                    report.empty_sku_rows.append(excel_row)
                else:
                    duplicate_candidates.setdefault(sku, []).append(excel_row)

                if name is None:
                    report.empty_name += 1

                if category is not None:
                    report.category_counts[category] = report.category_counts.get(category, 0) + 1

                price_value = self._cell_value(row, PRICE_COLUMN)
                if self._is_empty_cell(price_value):
                    report.empty_price += 1
                    report.empty_price_rows.append(excel_row)
                elif self._is_negative_number(price_value):
                    report.negative_price += 1

                quantity_value = self._cell_value(row, QUANTITY_COLUMN)
                if self._is_empty_cell(quantity_value):
                    report.empty_quantity += 1
                    report.empty_quantity_rows.append(excel_row)
                else:
                    if self._is_negative_number(quantity_value):
                        report.negative_quantity += 1
                    if self._is_formula_value(quantity_value):
                        report.formula_quantity += 1

                if shape is None:
                    report.empty_shape += 1
                if fixation is None:
                    report.empty_fixation += 1
                if cut is None:
                    report.empty_cut += 1
        finally:
            workbook.close()

        report.duplicate_sku_rows = {
            sku: rows for sku, rows in sorted(duplicate_candidates.items()) if len(rows) > 1
        }
        report.duplicate_sku = sum(len(rows) - 1 for rows in report.duplicate_sku_rows.values())
        report.category_counts = dict(sorted(report.category_counts.items()))
        return report

    @staticmethod
    def _cell_value(row: tuple[object, ...], index: int) -> object | None:
        """Return a cell value by column index when present."""
        if index >= len(row):
            return None

        cell = row[index]
        return None if cell is None else cell.value

    @staticmethod
    def _is_empty_row(row: tuple[object, ...]) -> bool:
        """Return whether a worksheet row has no meaningful values."""
        return all(CatalogAudit._is_empty_cell(cell.value if cell is not None else None) for cell in row)

    @staticmethod
    def _is_empty_cell(value: object | None) -> bool:
        """Return whether a cell value should be treated as empty."""
        if value is None:
            return True
        if isinstance(value, str):
            return not value.strip()
        return False

    @staticmethod
    def _string_value(value: object | None) -> str | None:
        """Return a stripped string value when present."""
        if CatalogAudit._is_empty_cell(value):
            return None
        return str(value).strip()

    @staticmethod
    def _is_negative_number(value: object | None) -> bool:
        """Return whether a cell contains a negative numeric value."""
        return isinstance(value, (int, float)) and value < 0

    @staticmethod
    def _is_formula_value(value: object | None) -> bool:
        """Return whether a cell contains a formula-like Excel value."""
        return isinstance(value, str) and value.startswith("=")


def main() -> int:
    """Run the catalog audit CLI and print the resulting summary."""
    if len(sys.argv) != 2:
        print("Usage:\n")
        print("python -m nas_parser.audit output/catalog.xlsx")
        return 1

    excel_file = Path(sys.argv[1])
    if not excel_file.is_file():
        print(f"File not found: {excel_file}")
        return 1

    report = CatalogAudit(excel_file).run()
    print(report.summary())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
