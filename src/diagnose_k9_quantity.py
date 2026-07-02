import openpyxl
from pathlib import Path

# Предполагаем, что K9Parser.QUANTITY_COLUMN = 7 (индекс 7 соответствует 8-й колонке Excel, т.е. 'H')
# Проверяем в nas_parser/parsers/k9.py: QUANTITY_COLUMN = 7
QUANTITY_COLUMN_INDEX = 7 

FILE_PATH = Path("input/K9.xlsx")

def diagnose_k9_quantity():
    if not FILE_PATH.exists():
        print(f"Ошибка: Файл {FILE_PATH} не найден.")
        return

    print(f"Анализ файла: {FILE_PATH}")
    print(f"Колонка Quantity (0-based index): {QUANTITY_COLUMN_INDEX} (колонка {chr(ord('A') + QUANTITY_COLUMN_INDEX)})")
    print("-" * 50)

    try:
        # Открываем книгу с data_only=False, чтобы получить формулы
        workbook_formulas = openpyxl.load_workbook(FILE_PATH, data_only=False, read_only=True)
        # Открываем книгу с data_only=True, чтобы получить вычисленные значения
        workbook_data = openpyxl.load_workbook(FILE_PATH, data_only=True, read_only=True)

        stats = {"numbers": 0, "formulas": 0, "strings": 0, "empty": 0, "other": 0}
        rows_processed = 0

        for sheet_name in workbook_formulas.sheetnames:
            sheet_f = workbook_formulas[sheet_name]
            sheet_d = workbook_data[sheet_name]
            
            print(f"\nЛист: {sheet_name}")
            
            for row_idx, row_f in enumerate(sheet_f.iter_rows(), start=1):
                if rows_processed >= 30:
                    break

                if QUANTITY_COLUMN_INDEX >= len(row_f):
                    continue

                cell_f = row_f[QUANTITY_COLUMN_INDEX]
                cell_d = sheet_d.cell(row=row_idx, column=QUANTITY_COLUMN_INDEX + 1) # openpyxl columns are 1-based

                # Пропускаем полностью пустые строки, но учитываем строки, где только Quantity пустое
                if all(c.value is None for c in row_f):
                    continue

                rows_processed += 1
                
                value = cell_f.value
                data_type = cell_f.data_type
                computed_value = cell_d.value # This will be the computed value if it's a formula

                print(f"  Строка {row_idx}, Ячейка {cell_f.coordinate}:")
                print(f"    cell.value (raw): {repr(value)}")
                print(f"    type(cell.value): {type(value)}")
                print(f"    cell.data_type: {data_type}")
                
                if data_type == 'f':
                    print(f"    Формула: {cell_f.formula}")
                    print(f"    Вычисленное значение (cached/data_only=True): {repr(computed_value)}")
                    stats["formulas"] += 1
                elif value is None:
                    print("    Пустое значение")
                    stats["empty"] += 1
                elif isinstance(value, (int, float)):
                    stats["numbers"] += 1
                elif isinstance(value, str):
                    stats["strings"] += 1
                else:
                    stats["other"] += 1
            
            if rows_processed >= 30:
                break

    except Exception as e:
        print(f"Произошла ошибка при чтении файла: {e}")
    finally:
        if 'workbook_formulas' in locals():
            workbook_formulas.close()
        if 'workbook_data' in locals():
            workbook_data.close()
    
    print("-" * 50)
    print("\nСтатистика по первым 30 непустым строкам в колонке Quantity:")
    for key, count in stats.items():
        print(f"- {key.capitalize()}: {count}")

if __name__ == "__main__":
    diagnose_k9_quantity()
