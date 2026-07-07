"""Split the consolidated inventory workbook into per-parser input files."""

from __future__ import annotations

from pathlib import Path
import sys
from typing import Sequence

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
INPUT_DIR = PROJECT_ROOT / "input"
SOURCE_PREFIX = "Остатки"

SHEET_TO_FILE: dict[str, str] = {
    "16 граней горячие": "16cut_hot.xlsx",
    "16 граней горячие АВ": "16cut_hot_AB.xlsx",
    "16 граней холодные": "16cut_non.xlsx",
    "16 граней холод АВ": "16cut_non_AB.xlsx",
    "12 граней горячие": "12cut_hot.xlsx",
    "12 граней горяч (mix)": "12cut_hot-mix.xlsx",
    "12 граней горячие АВ": "12cut_hot_AB.xlsx",
    "12 граней холодные": "12cut_non.xlsx",
    "12 граней холод (mix)": "12cut_non-mix.xlsx",
    "12 граней холод АВ": "12cut_non_AB.xlsx",
    "пришивные (K9)": "K9.xlsx",
    "K9 (премиум)": "K9_premium.xlsx",
}


def main(argv: Sequence[str] | None = None) -> int:
    """Split the inventory workbook into the input directory."""
    args = list(sys.argv[1:] if argv is None else argv)
    if len(args) > 1:
        _print_usage()
        return 1

    source_file = Path(args[0]) if args else _discover_source_file()
    if source_file is None:
        print(f'No inventory workbook found. Expected a file starting with "{SOURCE_PREFIX}".')
        return 1
    if not source_file.is_file():
        print(f"Inventory workbook not found: {source_file}")
        return 1

    INPUT_DIR.mkdir(parents=True, exist_ok=True)

    created = 0
    overwritten = 0
    skipped = 0
    unknown = 0
    processed_sheets: set[str] = set()

    workbook = load_workbook(source_file)
    try:
        for sheet_name in workbook.sheetnames:
            processed_sheets.add(sheet_name)
            target_file = SHEET_TO_FILE.get(sheet_name)
            if target_file is None:
                print(f"Unknown sheet:\n{sheet_name}")
                unknown += 1
                continue

            output_file = INPUT_DIR / target_file
            exists_before = output_file.exists()
            _save_single_sheet_copy(source_file, sheet_name, output_file)
            if exists_before:
                overwritten += 1
            else:
                created += 1
    finally:
        workbook.close()

    for sheet_name in SHEET_TO_FILE:
        if sheet_name not in processed_sheets:
            print(f"Skipped:\n{sheet_name}")
            skipped += 1

    print("Inventory split completed")
    print("Created:")
    print(created)
    print("Overwritten:")
    print(overwritten)
    print("Skipped:")
    print(skipped)
    print("Unknown:")
    print(unknown)
    return 0


def _discover_source_file() -> Path | None:
    """Find the single inventory workbook in the project root."""
    matches = sorted(
        path
        for path in PROJECT_ROOT.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx" and path.name.startswith(SOURCE_PREFIX)
    )
    if len(matches) == 1:
        return matches[0]

    return None


def _save_single_sheet_copy(source_file: Path, sheet_name: str, destination_file: Path) -> None:
    """Save a copy of the workbook containing only the requested sheet."""
    workbook = load_workbook(source_file)
    try:
        for current_sheet in list(workbook.sheetnames):
            if current_sheet != sheet_name:
                del workbook[current_sheet]

        workbook.active = 0
        workbook.save(destination_file)
    finally:
        workbook.close()


def _print_usage() -> None:
    """Print a short CLI usage message."""
    print("Usage:")
    print()
    print("python tools/split_inventory.py")
    print('python tools/split_inventory.py "<path_to_inventory.xlsx>"')


if __name__ == "__main__":
    raise SystemExit(main())
